import os
import pandas as pd
from openpyxl import load_workbook



# Buscar el archivo .xlsx más reciente que no sea base_codigos.xlsx
def buscar_excel_mas_reciente():
    archivos = [f for f in os.listdir() if f.endswith('.xlsx') and f != 'base_codigos.xlsx']
    if not archivos:
        return None
    archivos.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return archivos[0]

archivo_excel = buscar_excel_mas_reciente()
archivo_txt = "resultado.txt"  # Asegúrate que se llame así

if archivo_excel and os.path.exists(archivo_txt):
    print(f"[INFO] Excel encontrado: {archivo_excel}")
    print(f"[INFO] TXT encontrado: {archivo_txt}")

    # Leer líneas del txt
    with open(archivo_txt, "r", encoding="utf-8") as f:
        lineas = [line.strip() for line in f]

    # Cargar el Excel original con openpyxl
    wb = load_workbook(archivo_excel)
    if "Publicaciones" not in wb.sheetnames:
        print("[ERROR] La hoja 'Publicaciones' no existe en el Excel.")
    else:
        ws = wb["Publicaciones"]

        # Escribir cada línea en la columna F (columna 6), empezando desde fila 6
        for i, texto in enumerate(lineas, start=6):
            ws.cell(row=i, column=6, value=texto)

        # Crear carpeta temp si no existe
        carpeta_salida = "temp"
        if not os.path.exists(carpeta_salida):
            os.makedirs(carpeta_salida)

        nombre_sin_ext = os.path.splitext(archivo_excel)[0]
        nombre_resultado = f"resultado_{nombre_sin_ext}.xlsx"
        ruta_salida = os.path.join(carpeta_salida, nombre_resultado)
        wb.save(ruta_salida)
        print(f"[OK] Guardado en: {ruta_salida}")

        # Guardar nombre del archivo generado para que lo use la API de descarga
        with open(os.path.join(carpeta_salida, "ultimo_archivo.txt"), "w") as f:
            f.write(nombre_resultado)

else:
    print("[ERROR] No se encontró el Excel o el archivo resultado.txt.")